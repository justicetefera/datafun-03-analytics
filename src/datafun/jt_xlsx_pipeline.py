"""
jt_xlsx_pipeline.py - Custom XLSX ETVL pipeline.

Author: Justice Tefera
Date: 2026-05
"""

from pathlib import Path
from typing import Any, cast

import openpyxl
from openpyxl.cell.cell import Cell

# === E: EXTRACT ===


def extract_xlsx_column_strings(*, file_path: Path, column_letter: str) -> list[str]:
    """Extract non-empty strings from a column in an Excel file."""
    if not file_path.exists():
        raise FileNotFoundError(f"Input file not found: {file_path}")

    workbook = openpyxl.load_workbook(file_path)
    sheetnames = workbook.sheetnames
    sheet = workbook[sheetnames[0]]  # explicitly select the first sheet

    # === PRINT FEEDBACK FROM EXCEL ===
    print("\n=== FEEDBACK FROM EXCEL ===")
    for row in sheet.iter_rows(min_row=2, max_col=1, values_only=True):
        if row[0]:
            print(row[0])
    print("=== END FEEDBACK ===\n")

    values: list[str] = []
    for cell in sheet[column_letter]:
        cell = cast(Cell, cell)
        value = cell.value
        if isinstance(value, str) and value.strip():
            values.append(value)

    return values


# === T: TRANSFORM ===


def transform_count_word(*, values: list[str], word: str) -> dict[str, int]:
    """Count occurrences of a word and return count + rows scanned."""
    if not word:
        raise ValueError("Word to count cannot be empty.")

    target = word.lower()
    count = 0

    for text in values:
        count += text.lower().count(target)

    return {
        "count": count,
        "rows_scanned": len(values),
    }


# === V: VERIFY ===


def verify_results(*, results: dict[str, int]) -> None:
    """Verify count and rows scanned are valid."""
    if "count" not in results or "rows_scanned" not in results:
        raise KeyError("Missing required result keys.")

    if results["count"] < 0:
        raise ValueError("Count cannot be negative.")

    if results["rows_scanned"] <= 0:
        raise ValueError("Must scan at least one row.")


# === L: LOAD ===


def load_results_report(
    *, results: dict[str, int], out_path: Path, word: str, column_letter: str
) -> None:
    """Write results to a text file."""
    out_path.parent.mkdir(parents=True, exist_ok=True)

    with out_path.open("w", encoding="utf-8") as f:
        f.write("Justice's Custom XLSX Word Count Report\n")
        f.write("Analyzed Excel feedback data.\n\n")
        f.write(f"Word counted: {word}\n")
        f.write(f"Column scanned: {column_letter}\n")
        f.write(f"Rows scanned: {results['rows_scanned']}\n")
        f.write(f"Total occurrences: {results['count']}\n")


# === FULL PIPELINE ===


def run_xlsx_pipeline(*, raw_dir: Path, processed_dir: Path, logger: Any) -> None:
    """Run the full ETVL pipeline."""
    logger.info("JT XLSX PIPELINE: START")

    input_file = raw_dir / "jt_feedback.xlsx"
    output_file = processed_dir / "jt_xlsx_word_count.txt"

    column_letter = "A"
    word = "excellent"

    values = extract_xlsx_column_strings(
        file_path=input_file,
        column_letter=column_letter,
    )

    results = transform_count_word(values=values, word=word)

    verify_results(results=results)

    load_results_report(
        results=results,
        out_path=output_file,
        word=word,
        column_letter=column_letter,
    )

    logger.info("JT XLSX PIPELINE: wrote %s", output_file)
    logger.info("JT XLSX PIPELINE: END")
